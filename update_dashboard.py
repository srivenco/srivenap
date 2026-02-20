#!/usr/bin/env python3
"""
Sriven Market Dashboard — Weekly XLSX Updater
=============================================
Usage:
    python update_dashboard.py <xlsx_file> [label]

Example:
    python update_dashboard.py "OFFTAKES/AP RETAIL WISE OFFTAKE AS ON 16-02-2026.xlsx"
    python update_dashboard.py "OFFTAKES/AP RETAIL WISE OFFTAKE AS ON 16-02-2026.xlsx" "Feb 2026 (16 Feb)"

What it does:
    1. Reads the DEPO and SYNDICATE sheets from the weekly XLSX file
    2. Extracts depot-level and syndicate-level brand data
    3. Adds the new time period to the dashboard's RAW data
    4. Updates index.html in place (backs up the original)
    5. Adds a month pill button in the filter bar

Run this every time you receive a new weekly offtake file.
"""

import sys
import re
import json
import shutil
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ─── Brand column name → dashboard brand key ─────────────────────────────────
BRAND_MAP = {
    "DIRECTOR'S SPECIAL GOLD WHISKY":                    "DSG",
    "HAYWARDS FINE WHISKY":                              "HAYWARDS",
    "ROYAL STREET FINE GRAIN WHISKY":                    "ROYAL_STREET",
    "MANJEERA BLUE PREMIUM DELUXE WHISKY":               "MANJEERA",
    "OLD TAVERN DELUXE WHISKY":                          "OLD_TAVERN",
    "MCDOWELL'S NO 1 CELEBRATION PREMIUM XXX RUM":       "MCD_RUM",
    "OFFICERS CHOICE ELEGANT WHISKY":                    "OC_WHISKY",
    "SIRI'S CLASSIC BLUE FINEST WHISKY":                 "SIRIS_BLUE",
    "GRAYSON'S SILVER STRIPES ORIGINAL RESERVE WHISKY":  "GRAYSONS",
    "AC BLACK CLASSIC WHISKY":                           "AC_BLACK",
    "BAGPIPER GOLD RESERVE WHISKY":                      "BAGPIPER",
    "HYDERABAD BLUE RESERVE WHISKY":                     "HYD_BLUE",
    "8 PM GOLD BLEND OF SCOTCH & INDIAN GRAIN WHISKY":   "8PM",
    "GRAYSON'S KINGSWELL SELECT INDIAN BRANDY":          "KINGSWELL",
    "OFFICER'S CHOICE FINEST BRANDY":                    "OC_BRANDY",
    "NO 1 MCDOWELL'S SELECT GOLD INDIAN BRANDY":         "MCD_SELECT",
    "WHYTEHALL RARE PREMIUM BRANDY":                     "WHYTEHALL",
    "SIRI'S CLASSIC RED VSOP BRANDY":                    "SIRIS_RED",
    "ROYAL ARMS PREMIUM FRENCH BRANDY":                  "ROYAL_ARMS",
    "MCDOWELL'S VSOP FINE BLENDED BRANDY":               "MCD_VSOP",
    "NICOL'S BLACK & GOLD RARE PREMIUM FRENCH BRANDY VSOP": "NICOLS",
    "MELISSA PREMIUM BRANDY":                            "MELISSA",
    "SIRI'S FRENCH QUARTERS XO PREMIUM BRANDY":          "SIRIS_XO",
    "TI MANSION HOUSE PLATINUM FRENCH BRANDY":           "TI_MANSION",
}

# Segment aggregate columns in the XLSX → dashboard segment keys
SEG_MAP = {
    "150 WHISKY SEG":    "s140",
    "170 WHISKYSEG":     "s160w",
    "170 BRANDY SEG":    "s160b",
    "200 BRANDY SEG":    "s180",
    "160 TO 190 RUM SEG": "srum",
}

# Older file format (DE sheet) segment column names
SEG_MAP_OLD = {
    "140 W SEG":  "s140",
    "160 W SEG":  "s160w",
    "160 B SEG":  "s160b",
    "200 B SEG":  "s180",
}

# Depot code → standard name mapping (for older "DE" sheet files)
DEPOT_CODE_MAP = {
    "074-Anantapur":       "ANANTHAPUR",
    "101-Sri Satya Sai":   "SATYA SAI",
    "072-Kurnool":         "KURNOOL",
    "096-Nandyal":         "NANDYAL",
    "075-Kadapa":          "KADAPA-1",
    "095-Prodduturu":      "PRODDATUR",
    "073-Chittoor-I":      "CHITTOR-1",
    "089-Chittoor-II":     "CHITTOOR-2",
    "098-Chittoor-III":    "CHITTOR-3",
    "076-Nellore":         "NELLORE-1",
    "094-Nellore-II":      "NELLORE -2",
    "077-Prakasam-I":      "PRAKSAM-1",
    "091-Prakasam-II":     "PRAKSAM-2",
    "078-Guntur-I":        "GUNTUR-1",
    "079-Guntur-II":       "GUNTUR-2",
    "092-Guntur-III":      "GUNTUR-3",
    "080-Vijayawada-I":    "VIJAYAWADA-1",
    "081-Vijayawada-II":   "VIJAYAWADA-2",
    "097-Vijayawada-III":  "VIJAYAWADA-3",
    "083-East Godavari-I": "EAST GODAVRI -1",
    "084-East Godavari-II":"EAST GODAVARI-2",
    "093-East Godavari-III":"EAST GODAVARI-3",
    "082-West Godavari-I": "WEST GODAVARI-1",
    "090-West Godavari-II":"WEST GODAVARI -2",
    "100-Bhimavaram":      "WEST GODAVRI-3",
    "085-Vizag-I":         "VIZAG-1",
    "086-Vizag-II":        "VIZAG-2",
    "099-Vizag-III":       "VIZAG-3",
    "087-Vizianagaram":    "VIZAINAGARAM",
    "088-Srikakulam":      "SRIKAKULAM",
}

ALL_BRAND_KEYS = [
    "DSG","MANJEERA","OLD_TAVERN","ROYAL_STREET","HAYWARDS",
    "MCD_RUM","OC_WHISKY","SIRIS_BLUE","AC_BLACK","GRAYSONS",
    "HYD_BLUE","8PM","BAGPIPER","KINGSWELL","OC_BRANDY",
    "WHYTEHALL","SIRIS_RED","ROYAL_ARMS","MELISSA","MCD_SELECT",
    "MCD_VSOP","TI_MANSION","NICOLS","SIRIS_XO",
]


def normalise(s):
    """Strip leading/trailing spaces from a column header."""
    return str(s).strip() if s else ""


def parse_date_label_from_filename(xlsx_path):
    """
    Try to extract a human-readable label from the filename.
    'AP RETAIL WISE OFFTAKE AS ON 16-02-2026.xlsx' → 'Feb 2026 (16 Feb)'
    """
    name = Path(xlsx_path).stem
    match = re.search(r'(\d{1,2})[- ](\d{2})[- ](\d{4})', name)
    if match:
        day, month_num, year = match.groups()
        months = {
            "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
            "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
            "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
        }
        mon = months.get(month_num, month_num)
        return f"{mon} {year} ({day} {mon})"
    return name


def build_empty_row():
    row = {k: 0 for k in ALL_BRAND_KEYS}
    row.update({"s140": 0, "s160w": 0, "s160b": 0, "s180": 0, "srum": 0, "tot": 0})
    return row


def parse_depo_sheet(ws):
    """
    Parse the DEPO sheet.
    Returns:
        depots  – dict  depot_name → brand_values
        market  – dict  brand_key → total
    """
    headers = [normalise(c.value) for c in ws[1]]

    # Map col_index → brand_key or segment_key
    col_brand = {}
    col_seg   = {}
    for i, h in enumerate(headers):
        if h in BRAND_MAP:
            col_brand[i] = BRAND_MAP[h]
        elif h in SEG_MAP:
            col_seg[i] = SEG_MAP[h]

    depots = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        depot_raw = normalise(row[0]) if row[0] else ""
        outlet    = normalise(row[1]) if row[1] else ""

        if not depot_raw:
            continue

        # We want rows that are "XYZ Total" (aggregated per depot)
        if "Total" not in depot_raw:
            continue

        # Clean depot name: "ANANTHAPUR Total" → "ANANTHAPUR"
        depot = depot_raw.replace(" Total", "").strip()

        entry = build_empty_row()
        for i, bk in col_brand.items():
            entry[bk] += int(row[i] or 0)
        for i, sk in col_seg.items():
            entry[sk] += int(row[i] or 0)

        # Recompute total from the segment totals (more reliable)
        entry["tot"] = (entry["s140"] + entry["s160w"] +
                        entry["s160b"] + entry["s180"] + entry["srum"])

        if depot in depots:
            # Merge (shouldn't happen, but just in case)
            for k in entry:
                depots[depot][k] += entry[k]
        else:
            depots[depot] = entry

    # Build market totals
    market = build_empty_row()
    for dep_data in depots.values():
        for k in dep_data:
            market[k] += dep_data[k]

    return depots, market


def parse_de_sheet_old(ws):
    """
    Parse the older 'DE' sheet format where each row is a depot total.
    Depot names are in code format ('074-Anantapur') — mapped via DEPOT_CODE_MAP.
    """
    headers = [normalise(c.value) for c in ws[1]]

    col_brand = {}
    col_seg   = {}
    for i, h in enumerate(headers):
        if h in BRAND_MAP:
            col_brand[i] = BRAND_MAP[h]
        elif h in SEG_MAP_OLD:
            col_seg[i] = SEG_MAP_OLD[h]

    depots = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        depot_code = normalise(row[0]) if row[0] else ""
        if not depot_code or depot_code == "Grand Total":
            continue

        depot = DEPOT_CODE_MAP.get(depot_code, depot_code)

        entry = build_empty_row()
        for i, bk in col_brand.items():
            entry[bk] += int(row[i] or 0)
        for i, sk in col_seg.items():
            entry[sk] += int(row[i] or 0)

        entry["tot"] = (entry["s140"] + entry["s160w"] +
                        entry["s160b"] + entry["s180"])

        depots[depot] = entry

    market = build_empty_row()
    for dep_data in depots.values():
        for k in dep_data:
            market[k] += dep_data[k]

    return depots, market


def parse_syndicate_sheet(ws):
    """
    Parse the SYNDICATE sheet.
    Returns:
        syndicates – dict  syndicate_name → {depot, brand_values}
    """
    headers = [normalise(c.value) for c in ws[1]]

    col_brand = {}
    col_seg   = {}
    for i, h in enumerate(headers):
        if h in BRAND_MAP:
            col_brand[i] = BRAND_MAP[h]
        elif h in SEG_MAP:
            col_seg[i] = SEG_MAP[h]

    # depot col = 0, syndicate col = 1, outlet col = 2
    syndicates = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        depot_raw  = normalise(row[0]) if row[0] else ""
        syn_raw    = normalise(row[1]) if row[1] else ""

        if not syn_raw or "Total" not in syn_raw:
            continue

        # "KALYANI GROUP  Total" → "KALYANI GROUP"
        syn_name = re.sub(r'\s+Total\s*$', '', syn_raw).strip()
        depot    = depot_raw.strip()

        entry = build_empty_row()
        entry["depot"] = depot

        for i, bk in col_brand.items():
            entry[bk] += int(row[i] or 0)
        for i, sk in col_seg.items():
            entry[sk] += int(row[i] or 0)

        entry["tot"] = (entry["s140"] + entry["s160w"] +
                        entry["s160b"] + entry["s180"] + entry["srum"])

        if syn_name in syndicates:
            for k in entry:
                if k != "depot":
                    syndicates[syn_name][k] += entry[k]
        else:
            syndicates[syn_name] = entry

    return syndicates


def load_raw_from_html(html_path):
    """Extract the RAW JSON object from index.html."""
    content = Path(html_path).read_text(encoding="utf-8")
    match = re.search(r'const RAW\s*=\s*(\{.+?\});', content,
                      flags=re.DOTALL)
    if not match:
        raise ValueError("Could not find 'const RAW = {...}' in index.html")
    raw_json = match.group(1)
    return json.loads(raw_json), content, match.start(), match.end()


def write_raw_to_html(html_path, content, start, end, raw):
    """Replace the RAW JSON block in index.html with updated data."""
    new_raw = "const RAW=" + json.dumps(raw, ensure_ascii=False, separators=(',', ':')) + ";"
    new_content = content[:start] + new_raw + content[end:]
    Path(html_path).write_text(new_content, encoding="utf-8")


_MONTH_NUM = {
    'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
    'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12,
}

def is_weekly_label(label):
    """True if the label is a partial/weekly period (has a day in parentheses)."""
    return bool(re.search(r'\(\d+\s+\w+\)', label))

def label_sort_key(label):
    """
    Return a (year, month, day) tuple for chronological sorting.
    Monthly 'Jan 2026'        → (2026, 1, 31)   end-of-month so it sorts after weekly entries
    Weekly  'Feb 2026 (16 Feb)' → (2026, 2, 16)
    """
    day_match = re.search(r'\((\d+)\s+(\w+)\)', label)
    if day_match:
        day = int(day_match.group(1))
        mon = _MONTH_NUM.get(day_match.group(2), 1)
    else:
        day = 31
        for name, num in _MONTH_NUM.items():
            if name in label:
                mon = num
                break
        else:
            mon = 1
    year_match = re.search(r'(\d{4})', label)
    year = int(year_match.group(1)) if year_match else 2000
    return (year, mon, day)

def update_months_array(raw, label):
    """Insert label in the correct newest-first position."""
    months = raw.get("months", [])
    if label in months:
        return
    months = months + [label]
    months.sort(key=label_sort_key, reverse=True)   # newest first
    raw["months"] = months

def add_month_pill(content, label):
    """
    Inject a new month pill button into the filter bar in correct chronological
    position (newest first). Never changes which pill is currently selected.
    Weekly/partial periods get a 'pill wtd' CSS class for visual distinction.
    """
    # Detect existence via onclick pattern — avoids false-positives inside RAW JSON
    pill_onclick = f"setMonth('{label}'"
    if pill_onclick in content:
        print(f"  Month pill '{label}' already exists — skipping.")
        return content

    css_class = 'pill wtd' if is_weekly_label(label) else 'pill'
    new_pill = f'<div class="{css_class}" onclick="setMonth(\'{label}\',this)">{label}</div>\n  '

    # Find all existing pill buttons and their dates
    pill_re = re.compile(r'<div class="pill[^"]*" onclick="setMonth\(\'([^\']+)\'')
    existing = list(pill_re.finditer(content))

    if not existing:
        return content

    new_key = label_sort_key(label)
    for m in existing:
        if new_key > label_sort_key(m.group(1)):
            # Insert before this (older) pill
            content = content[:m.start()] + new_pill + content[m.start():]
            return content

    # Newer than nothing found → append after last pill's closing </div>
    last = existing[-1]
    end = content.find('</div>', last.start()) + len('</div>')
    content = content[:end] + '\n  ' + new_pill.rstrip() + content[end:]
    return content


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    label = sys.argv[2] if len(sys.argv) > 2 else parse_date_label_from_filename(xlsx_path)

    # Resolve paths relative to script location
    script_dir = Path(__file__).parent
    xlsx_full = Path(xlsx_path) if Path(xlsx_path).is_absolute() else script_dir / xlsx_path
    html_path = script_dir / "index.html"

    if not xlsx_full.exists():
        print(f"ERROR: File not found: {xlsx_full}")
        sys.exit(1)

    if not html_path.exists():
        print(f"ERROR: index.html not found at {html_path}")
        sys.exit(1)

    print(f"\nSriven Market Dashboard — Weekly Updater")
    print(f"{'─'*50}")
    print(f"  XLSX:  {xlsx_full.name}")
    print(f"  Label: {label}")
    print(f"  HTML:  {html_path}")
    print()

    # ── Step 1: Load workbook ──────────────────────────────────────────────────
    print("Reading XLSX...")
    wb = openpyxl.load_workbook(str(xlsx_full), data_only=True)

    if "DEPO" not in wb.sheetnames and "DE" not in wb.sheetnames:
        print(f"ERROR: Expected a 'DEPO' or 'DE' sheet. Found: {wb.sheetnames}")
        sys.exit(1)

    # ── Step 2: Parse sheets ───────────────────────────────────────────────────
    if "DEPO" in wb.sheetnames:
        print("Parsing DEPO sheet (new format)...")
        depots, market = parse_depo_sheet(wb["DEPO"])
    else:
        print("Parsing DE sheet (old format)...")
        depots, market = parse_de_sheet_old(wb["DE"])
    print(f"  → {len(depots)} depots found")
    print(f"  → Market total: {market['tot']:,} cases")

    syndicates = {}
    if "SYNDICATE" in wb.sheetnames:
        print("Parsing SYNDICATE sheet...")
        syndicates = parse_syndicate_sheet(wb["SYNDICATE"])
        print(f"  → {len(syndicates)} syndicates found")

    # ── Step 3: Load existing RAW data from HTML ───────────────────────────────
    print("Loading existing dashboard data...")
    raw, content, raw_start, raw_end = load_raw_from_html(html_path)

    # ── Step 4: Back up original ───────────────────────────────────────────────
    backup_path = html_path.with_suffix(f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
    shutil.copy2(html_path, backup_path)
    print(f"  Backup saved: {backup_path.name}")

    # ── Step 5: Inject new data period ────────────────────────────────────────
    print(f"Injecting period '{label}'...")

    if label in raw.get("depots", {}):
        print(f"  Period '{label}' already exists — will overwrite.")

    raw.setdefault("depots", {})[label] = depots
    raw.setdefault("market", {})[label] = market

    if syndicates:
        raw.setdefault("syndicates", {})[label] = syndicates

    update_months_array(raw, label)

    # ── Step 6: Write updated HTML ─────────────────────────────────────────────
    print("Writing updated index.html...")
    write_raw_to_html(html_path, content, raw_start, raw_end, raw)

    # ── Step 7: Add month pill to filter bar ───────────────────────────────────
    updated_content = Path(html_path).read_text(encoding="utf-8")
    updated_content = add_month_pill(updated_content, label)
    Path(html_path).write_text(updated_content, encoding="utf-8")

    # ── Step 8: Summary ───────────────────────────────────────────────────────
    print()
    print("✓ Dashboard updated successfully!")
    print()
    print("Our brands this period:")
    our_brands = {
        "DSG": "Director's Special Gold",
        "MCD_RUM": "McDowell's No.1 Rum",
        "GRAYSONS": "Grayson's Silver Stripes",
        "KINGSWELL": "Grayson's Kingswell",
        "MCD_SELECT": "McDowell's Select Gold",
        "MCD_VSOP": "McDowell's VSOP",
    }
    for bk, name in our_brands.items():
        val = market.get(bk, 0)
        print(f"  {name:<40} {val:>7,} cases")

    print()
    print("Open index.html in your browser to view the updated dashboard.")
    print()


if __name__ == "__main__":
    main()
